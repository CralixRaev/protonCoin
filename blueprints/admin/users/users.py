import io
from tempfile import NamedTemporaryFile

import flask
import openpyxl
import werkzeug
from flask import Blueprint, render_template, redirect, request, url_for, Request, Response
from flask_login import login_required
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import DimensionHolder, ColumnDimension
from werkzeug.datastructures import MultiDict

from blueprints.admin.users.forms.import_user import UserImportForm
from blueprints.admin.users.forms.user import UserForm
from db.models.group import GroupQuery
from db.models.user import UserQuery
from util import admin_required

users = Blueprint('users', __name__, template_folder='templates')


@users.route('/')
@login_required
@admin_required
def index():
    context = {
        'title': 'Пользователи',
        'users': UserQuery.get_all_users()
    }
    return render_template("users/users.html", **context)


@users.route('/create/', methods=['GET', 'POST'])
@login_required
@admin_required
def create_user():
    groups = GroupQuery.get_all_groups()
    group_list = [(-1, "Нет")] + [(group.id, group.name) for group in groups]
    form = UserForm()
    form.group_id.choices = group_list
    context = {
        'title': 'Создать пользователя',
        'form': form
    }
    if form.validate_on_submit():
        user, password = UserQuery.create_user(form.name.data, form.surname.data,
                                               form.patronymic.data,
                                               form.email.data if form.email.data else None,
                                               form.is_admin.data, form.is_teacher.data,
                                               form.group_id.data
                                               if form.group_id.data != -1 else None)
        flask.flash(f"Пользователь успешно создан. Его логин: {user.login}, пароль: {password}")
        return redirect(url_for('admin.users.index'))
    return render_template("users/user.html", **context)


    @users.route('/delete/')
    @login_required
    @admin_required
    def delete_user():
        user_id = request.args.get("id")
        user = UserQuery.get_user_by_id(user_id)
        UserQuery.delete_user(user)
        flask.flash(f"Пользователь ID: {user.id} - {user.full_name} успешно удалён")
        return redirect(url_for('admin.users.index'))


@users.route('/import/', methods=['GET', 'POST'])
@login_required
@admin_required
def import_users():
    groups = GroupQuery.get_all_groups()
    group_list = [(-1, "Нет")] + [(group.id, group.name) for group in groups]
    form = UserImportForm()
    form.group_id.choices = group_list
    context = {
        'title': 'Импортировать пользователя',
        'form': form
    }
    if form.validate_on_submit():
        table = form.file.data
        table = io.BytesIO(table.stream.read())
        wb_read = openpyxl.load_workbook(table)
        ws_read = wb_read.active
        wb_write = Workbook()
        ws_write = wb_write.active
        [ws_write.cell(1, i + 1, name) for i, name in enumerate(['ФИО', 'Логин', 'Пароль'])]
        for i, cells in enumerate(ws_read.iter_rows(2), start=2):
            full_name = [i.value for i in cells][0]
            split_name = full_name.split()
            surname, name, patronymic = split_name[0], split_name[1], ' '.join(split_name[2:])
            user, password = UserQuery.create_user(name, surname, patronymic if patronymic else None,
                                                   None, False, False,
                                                   form.group_id.data if form.group_id.data != -1 else None)
            ws_write.cell(i, 1, user.full_name)
            ws_write.cell(i, 2, user.login)
            ws_write.cell(i, 3, password)
        dim_holder = DimensionHolder(worksheet=ws_write)

        for col in range(ws_write.min_column, ws_write.max_column + 1):
            dim_holder[get_column_letter(col)] = ColumnDimension(ws_write, min=col, max=col,
                                                                 width=20)

        ws_write.column_dimensions = dim_holder
        with NamedTemporaryFile() as tmp:
            wb_write.save(tmp.name)
            output = io.BytesIO(tmp.read())
        flask.flash(f"Пользователи успешно созданы. Файл скачан к вам.")

        # forgive me please, this is high-load like code, just trust me
        def generate():
            for i in output:
                yield i

        return Response(generate(),
                        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    return render_template("users/import.html", **context)


@users.route('/edit/', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_user():
    user = UserQuery.get_user_by_id(request.args.get('id'))
    groups = GroupQuery.get_all_groups()
    group_list = [(-1, "Нет")] + [(group.id, group.name) for group in groups]
    form = UserForm()
    form.group_id.choices = group_list
    context = {
        'title': 'Редактировать пользователя',
        'form': form
    }
    if form.validate_on_submit():
        UserQuery.update_user(user, form.name.data, form.surname.data,
                              form.patronymic.data, form.email.data if form.email.data else None,
                              form.is_admin.data, form.is_teacher.data,
                              form.group_id.data if form.group_id.data != -1 else None)
        flask.flash(f"Пользователь успешно обновлен.")
        return redirect(url_for('admin.users.index'))
    model_data = MultiDict(user.__dict__.items())
    model_data['group_id'] = -1 if not model_data['group_id'] else model_data['group_id']
    form = UserForm(model_data)
    form.group_id.choices = group_list
    context['form'] = form
    return render_template("users/user.html", **context)


@users.route('/new_password/', methods=['GET', 'POST'])
@login_required
@admin_required
def new_password():
    password = UserQuery.new_password(request.args.get('id'))
    flask.flash(f"Новый пароль для пользователя: {password}")
    return redirect(url_for('admin.users.index'))
