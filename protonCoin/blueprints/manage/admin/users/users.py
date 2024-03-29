import io
from tempfile import NamedTemporaryFile

import flask
import openpyxl
from flask import Blueprint, render_template, redirect, url_for, Response, request
from flask_login import login_required
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.dimensions import DimensionHolder, ColumnDimension
from werkzeug.datastructures import MultiDict

from protonCoin.blueprints.manage.admin.users.forms.import_user import UserImportForm
from protonCoin.blueprints.manage.admin.users.forms.user import UserForm
from protonCoin.db.models.group import GroupQuery
from protonCoin.db.models.user import UserQuery
from protonCoin.util import admin_required

users = Blueprint(
    "users", __name__, template_folder="templates", static_folder="static"
)


@users.route("/")
@login_required
@admin_required
def index():
    context = {
        "title": "Пользователи",
    }
    return render_template("users/users.html", **context)


@users.route("/create/", methods=["GET", "POST"])
@login_required
@admin_required
def create_user():
    groups = GroupQuery.get_all_groups()
    group_list = [(-1, "Нет")] + [(group.id, group.name) for group in groups]
    form = UserForm()
    form.group_id.choices = group_list
    context = {"title": "Создать пользователя", "form": form}
    if form.validate_on_submit():
        user, password = UserQuery.create_user(
            form.name.data,
            form.surname.data,
            form.patronymic.data,
            form.email.data if form.email.data else None,
            form.is_admin.data,
            form.is_teacher.data,
            form.group_id.data if form.group_id.data != -1 else None,
        )
        flask.flash(
            f"Пользователь успешно создан. Его логин: {user.login}, пароль: {password}",
            "success",
        )
        return redirect(url_for(".index"))
    return render_template("users/user.html", **context)


@users.route("/import/", methods=["GET", "POST"])
@login_required
@admin_required
def import_users():
    groups = GroupQuery.get_all_groups()
    group_list = [(-1, "Нет")] + [(group.id, group.name) for group in groups]
    form = UserImportForm()
    form.group_id.choices = group_list
    context = {"title": "Импортировать пользователя", "form": form}
    if form.validate_on_submit():
        table = form.file.data
        table = io.BytesIO(table.stream.read())
        wb_read = openpyxl.load_workbook(table)
        ws_read = wb_read.active
        wb_write = Workbook()
        ws_write = wb_write.active
        [
            ws_write.cell(1, i + 1, name)
            for i, name in enumerate(["ФИО", "Логин", "Пароль"])
        ]
        for i, cells in enumerate(ws_read.iter_rows(2), start=2):
            full_name = next(i.value for i in cells)
            if full_name:
                split_name = full_name.split()
                surname, name, patronymic = (
                    split_name[0],
                    split_name[1],
                    " ".join(split_name[2:]),
                )
                user, password = UserQuery.create_user(
                    name,
                    surname,
                    patronymic if patronymic else None,
                    None,
                    False,
                    False,
                    form.group_id.data if form.group_id.data != -1 else None,
                )
                ws_write.cell(i, 1, user.full_name)
                ws_write.cell(i, 2, user.login)
                ws_write.cell(i, 3, password)
        dim_holder = DimensionHolder(worksheet=ws_write)

        for col in range(ws_write.min_column, ws_write.max_column + 1):
            dim_holder[get_column_letter(col)] = ColumnDimension(
                ws_write, min=col, max=col, width=20
            )

        ws_write.column_dimensions = dim_holder
        with NamedTemporaryFile() as tmp:
            wb_write.save(tmp.name)
            output = io.BytesIO(tmp.read())
        flask.flash("Пользователи успешно созданы. Файл скачан к вам.", "success")

        # forgive me please, this is high-load like code, just trust me
        def generate():
            yield from output

        return Response(
            generate(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    return render_template("users/import.html", **context)


@users.route("/edit/", methods=["GET", "POST"])
@login_required
@admin_required
def edit_user():
    user = UserQuery.get_user_by_id(request.args.get("id"))
    groups = GroupQuery.get_all_groups()
    group_list = [(-1, "Нет")] + [(group.id, group.name) for group in groups]
    form = UserForm()
    form.group_id.choices = group_list
    context = {"title": "Редактировать пользователя", "form": form}
    if form.validate_on_submit():
        UserQuery.update_user(
            user,
            form.name.data,
            form.surname.data,
            form.patronymic.data,
            form.email.data if form.email.data else None,
            form.is_admin.data,
            form.is_teacher.data,
            form.group_id.data if form.group_id.data != -1 else None,
        )
        flask.flash("Пользователь успешно обновлен.", "success")
        return redirect(url_for("manage.admin.users.index"))
    model_data = MultiDict(user.__dict__.items())
    model_data["group_id"] = (
        -1 if not model_data["group_id"] else model_data["group_id"]
    )
    form = UserForm(model_data)
    form.group_id.choices = group_list
    context["form"] = form
    return render_template("users/user.html", **context)
