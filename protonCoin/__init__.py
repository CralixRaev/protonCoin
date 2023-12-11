import logging
import os
from typing import Any

import click
import flask
from flask import Flask, abort, send_from_directory, Blueprint, render_template, request
from flask_login import LoginManager
from flask_mail import Mail
from flask_migrate import Migrate, Config
from flask_redis import Redis
from flask_uploads import configure_uploads, UploadSet
from protonCoin.db.__all_models import *  # noqa: F403
from protonCoin.blueprints.api.api import api_blueprint as api
from protonCoin.blueprints.landing.landing import landing
from protonCoin.blueprints.login.login import login

from protonCoin.blueprints.manage.manage import manage
from protonCoin.db.database import db
from protonCoin.db.models.group import GroupQuery
from protonCoin.db.models.transaction import TransactionQuery
from protonCoin.db.models.user import User, UserQuery
from protonCoin.uploads import avatars, gift_images, achievement_files
from protonCoin.util import teacher_or_admin_required


# ruff: noqa: PLC0415, PTH118, PTH119, PTH120, PLR0915, PLR2004, C901


def create_app():
    app = Flask(__name__)
    if app.debug:
        app.config.from_object("protonCoin.debug_config.DebugConfig")
    else:
        app.config.from_object(Config)
    _uploads = Blueprint("_uploads", "_uploads")

    # require admin for all not-proxied uploads
    @_uploads.route("/<setname>/<path:filename>")
    @teacher_or_admin_required
    def uploaded_file(setname: UploadSet, filename: str) -> Any:
        config = app.upload_set_config.get(setname)  # type: ignore
        if config is None:
            abort(404)
        return send_from_directory(config.destination, filename)

    @app.route("/webp_viewer/")
    @teacher_or_admin_required
    def webp_viewer():
        path = request.args.get("path")
        if not path:
            flask.abort(400)
        return render_template("webp_viewer.html", path=path)

    app.register_blueprint(_uploads, url_prefix="/uploads")
    app.register_blueprint(manage, url_prefix="/manage")
    app.register_blueprint(login, url_prefix="/login")
    app.register_blueprint(landing, url_prefix="/")
    app.register_blueprint(api, url_prefix="/api/v1/")

    login_manager = LoginManager()
    login_manager.init_app(app)
    login_manager.login_view = "login.index"
    login_manager.login_message = (
        "Пожалуйста, войдите, что бы получить доступ к этой странице"
    )

    @login_manager.user_loader
    def load_user(user_id):
        return User.query.filter(User.alternative_id == user_id).first()

    @app.cli.command("create_admin")
    @click.argument("name")
    @click.argument("surname")
    @click.argument("patronymic")
    def create_admin(name, surname, patronymic):
        user_object, password = UserQuery.create_user(
            name, surname, patronymic, is_admin=True
        )

        click.echo(f"Админ создан. Логин: {user_object.login}, пароль: {password}")

    @app.cli.command("reset_password")
    @click.argument("login")
    def reset_password(login):
        password = UserQuery.new_password(UserQuery.get_user_by_login(login).id)

        click.echo(f"Пароль изменён. Новый пароль: {password}")

    @app.cli.command("import_folders")
    @click.argument("folder")
    def import_folders(folder):
        from openpyxl.worksheet.dimensions import DimensionHolder
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.dimensions import ColumnDimension
        import openpyxl

        groups = os.listdir(os.path.join(folder, "for_import"))
        for group in groups:
            groupname = os.path.basename(group).split(".")[0]
            if groupname:
                stage, letter = groupname[0], groupname[1]
                wb_read = openpyxl.load_workbook(
                    os.path.join(folder, "for_import", group),
                    read_only=True,
                    data_only=True,
                )
                ws_read = wb_read.active
                wb_write = openpyxl.Workbook()
                ws_write = wb_write.active
                [
                    ws_write.cell(1, i + 1, name)
                    for i, name in enumerate(["ФИО", "Логин", "Пароль"])
                ]
                for row in ws_read.iter_rows(min_row=2):
                    full_name = row[0].value
                    if full_name:
                        split_name = full_name.split()
                        surname, name, patronymic = (
                            split_name[0],
                            split_name[1],
                            " ".join(split_name[2:]),
                        )
                        user, password = UserQuery.create_user(
                            name,
                            surname,
                            patronymic if patronymic else None,
                            None,
                            False,
                            False,
                            GroupQuery.get_group_by_stage_letter(stage, letter).id,
                        )
                        ws_write.append((user.full_name, user.login, password))
                dim_holder = DimensionHolder(worksheet=ws_write)

                for col in range(ws_write.min_column, ws_write.max_column + 1):
                    dim_holder[get_column_letter(col)] = ColumnDimension(
                        ws_write, min=col, max=col, width=20
                    )
                ws_write.column_dimensions = dim_holder

                wb_write.save(
                    os.path.join(folder, "imported", f"{groupname}-imported.xlsx")
                )
        click.echo("Ну, вроде импортировали!")

    @app.cli.command("import_teachers")
    @click.argument("file")
    def import_teachers(file):
        from openpyxl.worksheet.dimensions import DimensionHolder
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.dimensions import ColumnDimension
        import openpyxl

        wb_read = openpyxl.load_workbook(file, read_only=True, data_only=True)
        ws_read = wb_read.active
        wb_write = openpyxl.Workbook()
        ws_write = wb_write.active
        [
            ws_write.cell(1, i + 1, name)
            for i, name in enumerate(["ФИО", "Логин", "Пароль"])
        ]
        for row in ws_read.iter_rows(min_row=2):
            full_name = row[1].value
            if full_name:
                split_name = full_name.split()
                if len(row[0].value) == 2:
                    stage, letter = row[0].value[0], row[0].value[1]
                else:
                    stage, letter = row[0].value[0:2], row[0].value[2]
                surname, name, patronymic = (
                    split_name[0],
                    split_name[1],
                    " ".join(split_name[2:]),
                )
                user, password = UserQuery.create_user(
                    name,
                    surname,
                    patronymic if patronymic else None,
                    None,
                    False,
                    True,
                    GroupQuery.get_group_by_stage_letter(stage, letter).id,
                )
                ws_write.append((user.full_name, user.login, password))
        dim_holder = DimensionHolder(worksheet=ws_write)

        for col in range(ws_write.min_column, ws_write.max_column + 1):
            dim_holder[get_column_letter(col)] = ColumnDimension(
                ws_write, min=col, max=col, width=20
            )
        ws_write.column_dimensions = dim_holder

        wb_write.save("teachers.xlsx")
        click.echo("Ну, вро де импортировали!")

    @app.cli.command("vsoh_import")
    @click.argument("file")
    def vsoh_import(file):
        import openpyxl

        wb_read = openpyxl.load_workbook(file, read_only=True, data_only=True)
        for sheet in wb_read:
            for row in sheet.iter_rows(min_row=2):
                subject = row[5].value
                user_raw = row[1].value.split()
                surname = user_raw[0]
                name_first_letter = user_raw[1][0]
                patronymic_first_letter = user_raw[2][0] if len(user_raw) == 3 else ""
                user = User.query.filter(
                    User.surname == surname,
                    User.name.like(f"{name_first_letter}%"),
                    User.patronymic.like(f"{patronymic_first_letter}%"),
                ).all()
                if len(user) > 1 or len(user) < 1:
                    pass
                    # print(subject, surname, name_first_letter,
                    # patronymic_first_letter)
                else:
                    user = user[0]
                    TransactionQuery.create_accrual(
                        user.balance,
                        10,
                        f"За критерий (Участие во Всероссийской Олимпиаде"
                        f"Школьников) Школьный тур ({subject})"
                        f"(начислено автоматически)",
                    )
        click.echo("Ну, вроде импортировали!")

    db.init_app(app)
    _ = Migrate(app, db, render_as_batch=True, compare_type=True)

    _ = Redis(app)

    _ = Mail(app)

    configure_uploads(app, avatars)
    configure_uploads(app, gift_images)
    configure_uploads(app, achievement_files)

    # log only in production mode.
    if not app.debug:
        stream_handler = logging.StreamHandler()
        stream_handler.setLevel(logging.ERROR)
        app.logger.addHandler(stream_handler)

    # ensure what default "bank" balance is present
    # with app.app_context():
    # BalanceQuery.ensure_bank_balance()

    return app
